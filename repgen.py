# Import all the modules; import all the functions

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import pyexcel
import os
from time import sleep


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)



def convert_xls(xls_file):
    """
    xls_file : xls file.
    
    ------------
        takes an XLS file. 
    ------------
        creates a converted XLSX file.
    
    """

    pyexcel.save_book_as(file_name=xls_file,
                         dest_file_name=xls_file+"x")


def delete_file(file):
    """
    file : any file.
    
    ----------
        takes a file.
    ----------
        deletes it. 
    
    """

    os.remove(file)


def scrape_table(ws, content_range):
    """
  

    """
    scraped_info = []
    
    for row in ws[content_range]:
        for cell in row:
            scraped_info.append(cell.value)

    return scraped_info


def write_table(ws, rows, cols, row_range, col_range):
    """
    
    """
    # write the tables.

    for i, row in enumerate(ws[row_range]):
        for cell in row:
            cell.value = rows[i]
    

    j = 0
    for col in ws[col_range]:
        for cell in col:
            cell.value = cols[j]
            j += 1

def date_conversion(date_string):
    """
    Parameters
    ----------
    date_string : String containing a date. example:
        "110223"  or  "11/02/21"  or  "blahblahbvlah 110423"

    Returns
    -------
    DateTime Object.
    """
    # strip the symbols in between the dates if existant.

    rawdate = date_string.split(" ")[1]
    y = int('20'+rawdate[4:])
    m = int(rawdate[0:2])
    d = int(rawdate[2:4])

    return datetime(y, m, d)



def date_conversion_rev(dt):
    """
    Parameters
    -------
    datetime: datetime object.

    Returns
    -------
    String containing a datestring in the following format: "YYMMDD"
    """
    
    return dt.strftime("%m%d%y")
    
    
def move_sheet(wb, from_loc=None, to_loc=None):
    sheets=wb._sheets

    # if no from_loc given, assume last sheet
    if from_loc is None:
        from_loc = len(sheets) - 1

    #if no to_loc given, assume first
    if to_loc is None:
        to_loc = 0

    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)
    sleep(0.5)
    
    
def get_file_names():
    """
    Returns
    -------
    All filenames of this python file's current directory, excluding itself. 
    
    """

    current_dir = os.getcwd()

    file_list = os.listdir(current_dir)

    file_list.remove("growth_report.py")
   

    return file_list


def sort_by_date(file_list):
    """
    Parameters
    ----------
    file_list : List of filenames with dates on them.

    Returns
    -------
    Sorted List By Date.
    """

    sorted_files = []
    sorted_dates = []
    
    
    
    for file in file_list:
        if file == "LAST_GROWTH_REPORT.xlsx":
            pass
        else:
            sorted_dates.append(date_conversion(file))

    sorted_dates.sort()
    
    
    # FROM THE CURRENT DATETIME, GET LAST WEEK'S DATETIME
    for file in file_list:
        if file == "LAST_GROWTH_REPORT.xlsx":
            
            #needs present
            last_week = get_last_week(sorted_dates[0])
            
            #LAST WEEK STRING SHOULD REFLECT THE OTHER FILES' FORMATS : Y M D
            last_week_string = date_conversion_rev(last_week)
            
            newfilename = "report "+last_week_string+" lastweek.xlsx"
            os.rename("LAST_GROWTH_REPORT.xlsx", newfilename )
            
            #remove old filename and add new filename
            file_list.remove("LAST_GROWTH_REPORT.xlsx")
            
            current_dir = os.getcwd()
            file_list2 = os.listdir(current_dir)
            
            for file in file_list2:
                if file == newfilename:
                    file_list.append(newfilename)
                    
                else:
                    pass            
            
            #add the datetime to sorted_dates
            sorted_dates.insert(0, last_week)
            
    
    
    for dt in sorted_dates:
        for file in file_list:
            if date_conversion(file) == dt:

                sorted_files.append(file)
        
    return sorted_files, sorted_dates


def get_last_week(datetime):
    """
    Parameters
    ----------
    datetime : datetime object. 

    Returns
    -------
    datetime_lastweek : datetime object containing last week's date in
    relation to the received datetime.
    """
    return (datetime - timedelta(days=7))



def get_week_id(dt):
    return dt.strftime("%V")

def get_week_str(dt):
    return dt.strftime("%m/%d/%y")



